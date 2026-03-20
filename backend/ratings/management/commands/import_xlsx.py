from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from ratings.constants import PARAMETER_DEFS
from ratings.models import AirConditioner, ParameterValue

logger = logging.getLogger(__name__)

DEFAULT_FILENAME = "Рейтинг 2026-1.xlsx"


def _col_letter(cell: Any) -> str:
    """Return full column letter (handles AA, AB, etc.)."""
    idx = cell.column - 1
    if idx < 26:
        return chr(ord("A") + idx)
    return chr(ord("A") + idx // 26 - 1) + chr(ord("A") + idx % 26)


def _safe_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        return round(float(value), 2)
    except (ValueError, TypeError):
        return default


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


class Command(BaseCommand):
    help = "Импорт данных кондиционеров из xlsx-файла"

    def add_arguments(self, parser: Any) -> None:
        parser.add_argument(
            "file",
            nargs="?",
            default=None,
            help=f"Путь к xlsx-файлу (по умолчанию: {DEFAULT_FILENAME} в корне проекта)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Показать что будет импортировано, без записи в БД",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        file_path = self._resolve_path(options["file"])
        dry_run: bool = options["dry_run"]

        self.stdout.write(f"Загрузка из {file_path}...")
        if dry_run:
            self.stdout.write(self.style.WARNING("Режим dry-run: данные НЕ будут записаны"))

        wb = openpyxl.load_workbook(str(file_path), data_only=True)
        try:
            ws = wb.active
            if ws is None:
                raise CommandError("Файл не содержит активных листов")

            rows_data = self._parse_rows(ws)

            if dry_run:
                for row in rows_data:
                    self.stdout.write(f"  {row['rank']}. {row['brand']} — {row['model_name']} ({row['total_score']})")
                self.stdout.write(f"\nНайдено {len(rows_data)} моделей для импорта")
                return

            self._write_to_db(rows_data)
        finally:
            wb.close()

    def _resolve_path(self, file_arg: str | None) -> Path:
        if file_arg:
            path = Path(file_arg)
        else:
            path = Path(settings.BASE_DIR).parent / DEFAULT_FILENAME
        if not path.exists():
            raise CommandError(f"Файл не найден: {path}")
        return path

    def _parse_rows(self, ws: Any) -> list[dict[str, Any]]:
        rows_data: list[dict[str, Any]] = []
        skipped = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            cells: dict[str, Any] = {}
            for cell in row:
                cells[_col_letter(cell)] = cell.value

            rank = cells.get("A")
            if rank is None:
                skipped += 1
                continue

            brand = _safe_str(cells.get("B"))
            model_name = _safe_str(cells.get("C"))
            if not brand and not model_name:
                skipped += 1
                continue

            total_score = _safe_float(cells.get("AE"))
            if total_score == 0:
                logger.info("Row %d skipped: zero total score (%s %s)", row_idx, brand, model_name)
                skipped += 1
                continue

            try:
                rank_int = int(rank)
            except (ValueError, TypeError):
                logger.warning("Row %d: invalid rank '%s', skipping", row_idx, rank)
                skipped += 1
                continue

            parameters = []
            for pdef in PARAMETER_DEFS:
                parameters.append({
                    "parameter_name": pdef.name,
                    "raw_value": _safe_str(cells.get(pdef.value_col)),
                    "unit": pdef.unit,
                    "score": _safe_float(cells.get(pdef.score_col)),
                })

            rows_data.append({
                "rank": rank_int,
                "brand": brand,
                "model_name": model_name,
                "youtube_url": _safe_str(cells.get("D")),
                "rutube_url": _safe_str(cells.get("E")),
                "vk_url": _safe_str(cells.get("F")),
                "total_score": total_score,
                "parameters": parameters,
            })

        self.stdout.write(f"Распознано строк: {len(rows_data)}, пропущено: {skipped}")
        return rows_data

    def _write_to_db(self, rows_data: list[dict[str, Any]]) -> None:
        with transaction.atomic():
            AirConditioner.objects.all().delete()

            for row in rows_data:
                params = row.pop("parameters")
                ac = AirConditioner.objects.create(**row)
                ParameterValue.objects.bulk_create([
                    ParameterValue(air_conditioner=ac, **p) for p in params
                ])

        self.stdout.write(
            self.style.SUCCESS(f"Готово: импортировано {len(rows_data)} моделей")
        )
