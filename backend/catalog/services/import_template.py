"""Генерация XLSX-шаблона и выгрузки данных для команды import_v2."""

from __future__ import annotations

import re
from io import BytesIO

from django.db.models import Prefetch
from django.utils.text import slugify
import openpyxl
from openpyxl.styles import Font

from catalog.models import ACModel, ModelRawValue, ModelRegion
from methodology.models import Criterion, MethodologyVersion

# Порядок и имена колонок совпадают с import_v2 (management command).
FIXED_COLUMNS = [
    "brand",
    "model",
    "outer_unit",
    "series",
    "nominal_capacity",
    "equipment_type",
    "region",
    "youtube_url",
    "rutube_url",
    "vk_url",
    "compressor_model",
]


def _safe_filename_part(version: str) -> str:
    s = slugify(version.replace(".", "-"), allow_unicode=False) or "methodology"
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", s)[:80]


def _xlsx_cell(value):
    if value is None:
        return ""
    if isinstance(value, float) and value != value:  # NaN
        return ""
    return value


def _as_watts_capacity(value):
    if value is None:
        return ""
    try:
        num = float(value)
    except (ValueError, TypeError):
        return value
    # Stored in watts; legacy rows in kW are converted for export.
    if 0 < num < 100:
        num = num * 1000.0
    return int(num) if num.is_integer() else round(num, 3)


def generate_import_template_xlsx() -> tuple[bytes, str]:
    """
    Собирает книгу Excel: заголовки + справочник критериев + строки по всем моделям в БД.

    Returns:
        (содержимое .xlsx, имя файла для Content-Disposition)

    Raises:
        ValueError: если нет активной методики.
    """
    methodology = MethodologyVersion.objects.filter(is_active=True).first()
    if methodology is None:
        raise ValueError("Нет активной методики — шаблон недоступен.")

    criteria = list(
        Criterion.objects.filter(
            methodology=methodology,
            is_active=True,
        ).order_by("display_order", "code"),
    )
    crit_ids = [c.pk for c in criteria]
    code_list = [c.code for c in criteria]
    headers = FIXED_COLUMNS + code_list

    rv_prefetch = Prefetch(
        "raw_values",
        queryset=ModelRawValue.objects.filter(criterion_id__in=crit_ids).select_related("criterion"),
    )

    ac_models = (
        ACModel.objects.select_related("brand", "equipment_type")
        .prefetch_related("regions", rv_prefetch)
        .order_by("brand__name", "inner_unit")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Импорт"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    allowed_regions = {choice.value for choice in ModelRegion.RegionCode}
    for ac in ac_models:
        region_codes = sorted(
            {r for r in ac.regions.values_list("region_code", flat=True) if r in allowed_regions},
        )
        region_cell = ",".join(region_codes) if region_codes else "ru"

        raw_by_code: dict[str, str] = {}
        compressor_model = ""
        for rv in ac.raw_values.all():
            code = rv.criterion.code
            if code in code_list:
                raw_by_code[code] = rv.raw_value or ""
            if code == "compressor_power":
                compressor_model = (rv.compressor_model or "").strip()

        row_fixed = [
            ac.brand.name,
            ac.inner_unit,
            ac.outer_unit or "",
            ac.series or "",
            _as_watts_capacity(ac.nominal_capacity),
            ac.equipment_type.name if ac.equipment_type_id else "",
            region_cell,
            ac.youtube_url or "",
            ac.rutube_url or "",
            ac.vk_url or "",
            compressor_model,
        ]
        row_criteria = [raw_by_code.get(code, "") for code in code_list]
        ws.append(row_fixed + row_criteria)

    ws2 = wb.create_sheet("Критерии", 1)
    ws2.append(["code", "name_ru", "unit", "weight"])
    for c in criteria:
        ws2.append([c.code, c.name_ru, c.unit or "", float(c.weight)])

    for column_cells in ws2.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws2.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

    bio = BytesIO()
    wb.save(bio)
    body = bio.getvalue()
    fname = f"models_export_{_safe_filename_part(methodology.version)}.xlsx"
    return body, fname
