"""Импорт моделей из CSV/XLS/XLSX."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

import openpyxl
import xlrd
from django.db import transaction

from brands.models import Brand
from catalog.models import ACModel, EquipmentType, ModelRawValue, ModelRegion
from catalog.sync_brand_age import sync_brand_age_for_model
from methodology.models import Criterion, MethodologyVersion
from scoring.engine import update_model_total_index


def _normalize_model_name(val: Any) -> str:
    return str(val or "").strip().upper()


def _normalize_decimal_string(val: Any) -> str:
    return str(val).strip().replace(",", ".")


def _safe_float(val: Any) -> float | None:
    if val is None:
        return None
    try:
        num = float(_normalize_decimal_string(val))
        # nominal_capacity is stored in watts; support legacy kW inputs.
        if 0 < num < 100:
            num = num * 1000.0
        return num
    except (ValueError, TypeError):
        return None


def _is_empty(val: Any) -> bool:
    return val is None or str(val).strip() == ""


def _prepare_criterion_value(criterion: Criterion, val: Any) -> tuple[str | None, float | None]:
    """
    Возвращает (raw_value, numeric_value) или (None, None), если значение пустое
    либо конфликтует с типом критерия.
    """
    if _is_empty(val):
        return None, None

    normalized = _normalize_decimal_string(val)
    vtype = criterion.value_type
    numeric_types = {
        Criterion.ValueType.NUMERIC,
        Criterion.ValueType.LAB,
        Criterion.ValueType.FORMULA,
        Criterion.ValueType.BRAND_AGE,
        Criterion.ValueType.FALLBACK,
    }

    if vtype in numeric_types:
        try:
            numeric = float(normalized)
        except (ValueError, TypeError):
            return None, None
        # compressor_power is stored in watts; support legacy kW inputs.
        if criterion.code == "compressor_power" and 0 < numeric < 100:
            numeric = numeric * 1000.0
            normalized = f"{numeric:g}"
        return normalized, numeric

    if vtype == Criterion.ValueType.BINARY:
        token = normalized.lower()
        valid_tokens = {
            "1", "0",
            "yes", "no",
            "true", "false",
            "да", "нет",
        }
        if token not in valid_tokens:
            return None, None
        return normalized, None

    return str(val).strip(), None


def _read_csv(path: Path) -> list[dict]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _read_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    if ws is None:
        return []
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        data = {headers[i]: (row[i].value if i < len(row) else "") for i in range(len(headers))}
        rows.append(data)
    wb.close()
    return rows


def _read_xls(path: Path) -> list[dict]:
    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)
    if ws.nrows == 0:
        return []
    headers = [str(ws.cell_value(0, c) or "").strip() for c in range(ws.ncols)]
    rows: list[dict] = []
    for r in range(1, ws.nrows):
        data = {headers[c]: ws.cell_value(r, c) if c < ws.ncols else "" for c in range(len(headers))}
        rows.append(data)
    return rows


def _read_rows(path: Path) -> list[dict]:
    ext = path.suffix.lower()
    if ext == ".csv":
        return _read_csv(path)
    if ext == ".xls":
        return _read_xls(path)
    if ext == ".xlsx":
        return _read_xlsx(path)
    raise ValueError(f"Неподдерживаемый формат: {path.suffix}")


def find_existing_models_in_rows(rows: list[dict]) -> list[str]:
    """Вернёт список уникальных одноимённых моделей, уже существующих в БД."""
    existing_labels: list[str] = []
    seen: set[tuple[str, str]] = set()
    for row in rows:
        brand_name = str(row.get("brand", "")).strip()
        model_name = _normalize_model_name(row.get("model", ""))
        if not brand_name or not model_name:
            continue
        key = (brand_name, model_name)
        if key in seen:
            continue
        seen.add(key)
        if ACModel.objects.filter(brand__name=brand_name, inner_unit=model_name).exists():
            existing_labels.append(f"{brand_name} {model_name}")
    return existing_labels


def find_existing_models_in_file(path: Path) -> list[str]:
    return find_existing_models_in_rows(_read_rows(path))


def import_models_from_file(path: Path, *, publish: bool = False) -> tuple[int, list[str]]:
    methodology = MethodologyVersion.objects.filter(is_active=True).first()
    if not methodology:
        raise ValueError("Нет активной методики")

    criteria = {
        c.code: c for c in Criterion.objects.filter(
            methodology=methodology, is_active=True,
        )
    }

    rows = _read_rows(path)

    status = ACModel.PublishStatus.PUBLISHED if publish else ACModel.PublishStatus.DRAFT
    errors: list[str] = []
    imported = 0

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            brand_name = str(row.get("brand", "")).strip()
            model_name = _normalize_model_name(row.get("model", ""))
            if not brand_name or not model_name:
                errors.append(f"Строка {idx}: пустой бренд или модель")
                continue

            # Если бренд новый — создаётся с пустыми origin_class/sales_start_year_ru.
            brand, _ = Brand.objects.get_or_create(name=brand_name)
            ac = ACModel.objects.filter(brand=brand, inner_unit=model_name).first()
            if ac is None:
                eq_type, _ = EquipmentType.objects.get_or_create(
                    name=str(row.get("equipment_type", "Настенная сплит-система")),
                )
                ac = ACModel.objects.create(
                    brand=brand,
                    inner_unit=model_name,
                    outer_unit=str(row.get("outer_unit", "")),
                    series=str(row.get("series", "")),
                    nominal_capacity=_safe_float(row.get("nominal_capacity")),
                    equipment_type=eq_type,
                    publish_status=status,
                    youtube_url=str(row.get("youtube_url", "")),
                    rutube_url=str(row.get("rutube_url", "")),
                    vk_url=str(row.get("vk_url", "")),
                )
                imported += 1
            else:
                errors.append(
                    f"Строка {idx}: модель уже существует ({brand_name} {model_name}) — обновлены только критерии.",
                )

            region_raw = str(row.get("region", "ru") or "").strip()
            tokens = (
                [x.strip() for x in region_raw.split(",") if x.strip()]
                if region_raw
                else []
            )
            allowed = {c.value for c in ModelRegion.RegionCode}
            to_add = [c for c in tokens if c in allowed]
            if not to_add:
                to_add = ["ru"]
            for code in to_add:
                ModelRegion.objects.get_or_create(model=ac, region_code=code)

            for code, criterion in criteria.items():
                raw_value, numeric = _prepare_criterion_value(criterion, row.get(code, ""))
                if raw_value is None:
                    continue
                compressor_model = ""
                if code == "compressor_power":
                    compressor_model = str(row.get("compressor_model", "") or "").strip()
                ModelRawValue.objects.update_or_create(
                    model=ac,
                    criterion=criterion,
                    defaults={
                        "raw_value": raw_value,
                        "compressor_model": compressor_model,
                        "numeric_value": numeric,
                        "source": f"Импорт из {path.name}",
                    },
                )

            sync_brand_age_for_model(ac)
            update_model_total_index(ac)

    return imported, errors
