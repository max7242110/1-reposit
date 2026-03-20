"""Import data from CSV or XLSX with validation against active methodology (TZ section 23)."""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from brands.models import Brand
from catalog.models import ACModel, EquipmentType, ModelRawValue, ModelRegion
from methodology.models import Criterion, MethodologyVersion

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = "Импорт моделей из CSV/XLSX с валидацией по активной методике"

    def add_arguments(self, parser: Any) -> None:
        parser.add_argument("file", help="Путь к файлу (CSV или XLSX)")
        parser.add_argument(
            "--publish", action="store_true",
            help="Сразу публиковать (по умолчанию — черновик)",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        path = Path(options["file"])
        if not path.exists():
            raise CommandError(f"Файл не найден: {path}")

        methodology = MethodologyVersion.objects.filter(is_active=True).first()
        if not methodology:
            raise CommandError("Нет активной методики")

        criteria = {
            c.code: c for c in Criterion.objects.filter(
                methodology=methodology, is_active=True,
            )
        }

        if path.suffix.lower() == ".csv":
            rows = self._read_csv(path)
        elif path.suffix.lower() == ".xlsx":
            rows = self._read_xlsx(path)
        else:
            raise CommandError(f"Неподдерживаемый формат: {path.suffix}")

        publish = options["publish"]
        status = ACModel.PublishStatus.PUBLISHED if publish else ACModel.PublishStatus.DRAFT

        errors = []
        imported = 0

        with transaction.atomic():
            for idx, row in enumerate(rows, start=2):
                brand_name = row.get("brand", "").strip()
                model_name = row.get("model", "").strip()
                if not brand_name or not model_name:
                    errors.append(f"Строка {idx}: пустой бренд или модель")
                    continue

                brand, _ = Brand.objects.get_or_create(name=brand_name)
                eq_type, _ = EquipmentType.objects.get_or_create(
                    name=row.get("equipment_type", "Настенная сплит-система"),
                )

                ac = ACModel.objects.create(
                    brand=brand,
                    inner_unit=model_name,
                    outer_unit=row.get("outer_unit", ""),
                    series=row.get("series", ""),
                    nominal_capacity=self._safe_float(row.get("nominal_capacity")),
                    equipment_type=eq_type,
                    publish_status=status,
                    youtube_url=row.get("youtube_url", ""),
                    rutube_url=row.get("rutube_url", ""),
                    vk_url=row.get("vk_url", ""),
                )

                region = row.get("region", "ru")
                if region:
                    ModelRegion.objects.create(model=ac, region_code=region)

                for code, criterion in criteria.items():
                    val = row.get(code, "")
                    if val:
                        numeric = None
                        try:
                            numeric = float(val)
                        except (ValueError, TypeError):
                            pass
                        ModelRawValue.objects.create(
                            model=ac, criterion=criterion,
                            raw_value=str(val), numeric_value=numeric,
                            source=f"Импорт из {path.name}",
                        )
                imported += 1

        for err in errors:
            self.stderr.write(self.style.WARNING(err))

        self.stdout.write(self.style.SUCCESS(
            f"Импортировано {imported} моделей, ошибок: {len(errors)}"
        ))

    def _safe_float(self, val: Any) -> float | None:
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _read_csv(self, path: Path) -> list[dict]:
        with open(path, newline="", encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))

    def _read_xlsx(self, path: Path) -> list[dict]:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        ws = wb.active
        if ws is None:
            return []
        headers = [str(c.value or "").strip() for c in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            data = {headers[i]: (row[i].value if i < len(row) else "") for i in range(len(headers))}
            rows.append(data)
        wb.close()
        return rows
